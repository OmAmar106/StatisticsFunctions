def TableMaker():
    import xlsxwriter
    row = 0
    column = 0
    workbook = xlsxwriter.Workbook("Statistics.xlsx")
    worksheet = workbook.add_worksheet()
    a = int(input("Enter number of records : "))
    b = int(input("Enter number of columns : "))
    for i in range (a) :
        L = []
        for j in range (b) :
            c = j+1
            d = i+1
            st = "For Record "+str(d)+" , column "+str(c)+" , enter element"
            print(st)
            x = eval(input("Enter Element : "))
            L.append(x)
        for item in L :
            worksheet.write(row,column,item)
            if (column+1) < b:
                column = column + 1
            else :
                pass
        row = row + 1
        column = 0 
    workbook.close()
def TableReader():
    global L 
    import openpyxl
    L = []
    wb_obj = openpyxl.load_workbook("Statistics.xlsx")
    sheet_obj = wb_obj.active
    column = sheet_obj.max_column
    row = sheet_obj.max_row
    for i in range(1,row+1):
        L1 = []
        for j in range(1,column+1):
            cell_obj = sheet_obj.cell(row=i,column=j)
            L1.append(cell_obj.value)
        L.append(L1)
def Mean(): 
    c = f = 0 
    TableReader()
    a = int(input("Enter column Fi : "))
    a = a-1
    b = int(input("Enter column xi : "))
    b = b-1
    for item in L :
        d = item[a]
        g = item[b]
        e = d*g
        c = c+e
        f = f+d
        l = c/f
    print("Mean is : ",l)
def LinearPercentile(): 
    import math
    a = int(input("Enter percentile : "))
    b = int(input("Enter Sample Size : "))
    if math.ceil(a*b/100) == (a*b/100):
        c = (a*b/100)
        print("The avergae of ",c,"and",(c+1)," smallest value is at ",a,"th percentile")
    else:  
        c = math.ceil(a*b/100)
        print("The ",c," smallest value is at ",a,"th percentile")
def TablePercentile():
    import math
    TableReader()
    L2 = []
    a = int(input("Enter Column number with values : "))
    a = a-1
    for i in L :
        L2.append(i[a])
    L2.sort()
    a = int(input("Enter percentile(not 100) : "))
    b = len(L2)
    if math.ceil(a*b/100) == (a*b/100):
        c = (a*b/100)
        c = c-1
        print("The value is : ",(L2[c]+L2[c+1])/2)
    else:  
        c = math.ceil(a*b/100)
        c = c-1
        print("The value is : ",L2[c])
def SampleVariance():
    global k
    TableReader()
    it = int(input("Enter column no. of xi : "))
    it = it - 1
    c = 0
    h = 0
    for item in L :
        g = item[it]
        c = c+g
        a = len(L)
    m = c/a
    for item in L :
        g = item[it]
        l = (g-m)*(g-m)
        h = h+l
    k = h/(a-1)
    print("Sample Variance is : ",k)
def SampleStandardDeviation():
    SampleVariance()
    g = k**(1/2)
    print("Sample Standard Deviaiton is : ",g)
def SampleCorrelationCoeffecient():  #yet to do(i think not in syllabus)
    pass
def ExpectedValue():
    TableReader()
    global c
    c = 0
    a = int(input("Enter column no. of xi : "))
    a = a-1
    b = int(input("Enter column no. of P[xi](in fraction) : "))
    b = b-1
    for item in L :
        l = item[a] * item[b]
        c = c+l
    print("The Expected Value is : ",c)
def Variance():
    global f
    TableReader()
    e = 0
    a = int(input("Enter column no. of xi : "))
    a = a-1
    b = int(input("Enter column no. of P[xi](in fraction) : "))
    b = b-1
    for item in L :
        v = item[a]*item[a]
        e = e + (v*(item[b]))
    ExpectedValue()
    u = c*c
    f = e-u
    print("The Variance is : ",f)
def StandardDeviation():
    Variance()
    k = f**(1/2)
    print("The Standard Deviation is : ",k)
def BinomialRandomVariable():
    a = int(input("Enter numerator of probablity : "))
    b = int(input("Enter denominator : "))
    p = a/b
    q = 1-p
    i = int(input("Enter no. of success : "))
    n = int(input("Total no. of trials : "))
    import math
    f = math.factorial(n)/(math.factorial(i)*math.factorial(n-i))
    l = p**i
    o = q**(n-i)
    print("The Binomial Random Variable is : ",o*f*l)
    print("E[x] : ",n*p)
    print("Var[x] : ",n*p*q)
def StandardNormalProbablities():
    a = int(input("Value should be greater then(Value or NA) : "))
    if a == "NA":
        from PIL import Image
        img = Image.open(r"C:\Users\amarp\Documents\Code\IITM Statistics\Standard Normal Probablities(Z).png")
        img.show()
        l = float(input("Enter value of P(Z<.. ) : "))
        b = float(input("Enter Mean : "))
        c = float(input("Enter Standard Deviation : "))
        print("Value greater than : ",l*c + b)
    else:
        b = float(input("Enter Mean : "))
        c = float(input("Enter Standard Deviation : "))
        Z = (a-b)/c
        print("Z = ",Z)
        from PIL import Image
        img = Image.open(r"C:\Users\test\Documents\Code\Statistics Question Solver\Standard Normal Probablities(Z).png")
        img.show()
def StandardNormalProbablityCLT():
    d = int(input("Enter Mean : "))
    f = int(input("Etner Total number : "))
    l = d/(f**(1/2))
    a = int(input("Value should be greater then(Value) : "))
    Z = (a-d)/l
    print("Z = ",Z)
    from PIL import Image
    img = Image.open(r"C:\Users\test\Documents\Code\Statistics Question Solver\Standard Normal Probablities(Z).png")
    img.show()
def StandardDeviationNat():
    p = float(input("Enter Probablity of success : "))
    n = int(input("Enter total number of entries : "))
    Sd = (p*(1-p)/n)**1/2
    print("SD(xmean) = ",Sd)
    print("This is for equally occuring events , pg 339 in ross s")
print("1. Create Table")
print("2. Mean")
print("3. Linear Percentile")
print("4. Tabe Percentile")
print("5. Sample Variance")
print("6. Sample Standard Deviation")
print("7. Sample Correlation Coeffecienct")
print("8. Expected Value")
print("9. Variance")
print("10. Standard Deviation")
print("11. Binomial Random Variable")
print("12. Standard Normal Probablity")
print("13. Standard Normal Probablity using Central limit theorem")
print("14. Standard Deviaiton Natural")
print("15. EXIT")
while True:
    check = int(input("Enter number from 1 to 15 : "))
    if check == 1 :
        TableMaker()
    elif check == 2:
        Mean()
    elif check == 3:
        LinearPercentile()
    elif check == 4:
        TablePercentile()
    elif check == 5:
        SampleVariance()
    elif check == 6:
        SampleStandardDeviation()
    elif check == 7:
        SampleCorrelationCoeffecient()
    elif check == 8:
        ExpectedValue()
    elif check == 9:
        Variance()
    elif check == 10:
        StandardDeviation()
    elif check == 11:
        BinomialRandomVariable()
    elif check == 12:
        StandardNormalProbablities()
    elif check == 13:
        StandardNormalProbablityCLT()
    elif check == 14:
        StandardDeviationNat()
    else: 
        break
    
    
    
    
    
    
        
    

        
        
        

        
            
        

            
        
